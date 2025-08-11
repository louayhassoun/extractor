from flask import Flask, render_template, request, send_file, redirect, url_for
import pandas as pd
import re
import os
from openpyxl import load_workbook
from openpyxl.styles import Alignment

app = Flask(__name__)
UPLOAD_FOLDER = "uploads"
OUTPUT_FILE = "output_with_emails.xlsx"
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

EMAIL_REGEX = r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}"

def find_emails(text):
    if pd.isna(text):
        return []
    return re.findall(EMAIL_REGEX, str(text))

def process_excel(filepath, description_column_name="Description"):
    # Read excel (first sheet)
    df = pd.read_excel(filepath, dtype=str)  # read as strings to avoid mixed types
    # Ensure column exists (case-insensitive search)
    cols_lower = {c.lower(): c for c in df.columns}
    if description_column_name.lower() not in cols_lower:
        # try to guess description column if not exact
        # prefer columns containing 'desc' or 'description' or 'info'
        guess = None
        for k,v in cols_lower.items():
            if "organisation" in k or "info" in k or "description" in k:
                guess = v
                break
        if guess:
            description_column = guess
        else:
            # fallback to first text column
            description_column = df.columns[0]
    else:
        description_column = cols_lower[description_column_name.lower()]

    # Build Emails column: multiple emails joined by newline
    emails_list = []
    for idx, row in df.iterrows():
        text = row.get(description_column, "")
        emails = find_emails(text)
        if emails:
            # remove duplicates while preserving order
            seen = []
            for e in emails:
                if e not in seen:
                    seen.append(e)
            emails_list.append("\n".join(seen))
        else:
            emails_list.append("")  # empty cell if none found

    df["Emails"] = emails_list

    # Save to Excel
    df.to_excel(OUTPUT_FILE, index=False)

    # Use openpyxl to set wrap_text for Emails column and auto row height-ish
    try:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb.active
        # Find column index for "Emails"
        col_index = None
        for col in range(1, ws.max_column + 1):
            if ws.cell(row=1, column=col).value == "Emails":
                col_index = col
                break
        if col_index:
            # set wrap_text for all cells in that column
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_index)
                if cell.value is None:
                    continue
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            # Optionally set column width
            ws.column_dimensions[ws.cell(row=1, column=col_index).column_letter].width = 30
        wb.save(OUTPUT_FILE)
    except Exception as e:
        # if openpyxl fails, we still have the file saved by pandas
        print("Warning: could not set Excel cell wrap formatting:", e)

    return OUTPUT_FILE, description_column

# -------------------
# Routes
# -------------------
@app.route("/", methods=["GET", "POST"])
def index():
    if request.method == "POST":
        uploaded = request.files.get("excel_file")
        desc_col_input = request.form.get("description_column", "").strip()
        if not uploaded or uploaded.filename == "":
            return render_template("index.html", error="Please upload an Excel file (.xlsx or .xls).")
        filename = uploaded.filename
        save_path = os.path.join(UPLOAD_FOLDER, filename)
        uploaded.save(save_path)

        # default description column
        desc_col = desc_col_input if desc_col_input else "Description"

        try:
            output_path, used_desc_col = process_excel(save_path, description_column_name=desc_col)
        except Exception as e:
            return render_template("index.html", error=f"Error processing file: {e}")

        # read preview dataframe (first N rows)
        preview_df = pd.read_excel(output_path)
        # show only first 200 rows to avoid huge pages
        preview_html = preview_df.head(200).to_html(classes="table table-striped", index=False, escape=False)

        return render_template("preview.html", tables=[preview_html], filename=os.path.basename(output_path),
                               used_desc_col=used_desc_col)

    return render_template("index.html")

@app.route("/download/<filename>")
def download(filename):
    path = os.path.join(".", filename)
    if os.path.exists(path):
        return send_file(path, as_attachment=True)
    return redirect(url_for('index'))

# Optional reset route to delete last output
@app.route("/reset")
def reset_output():
    try:
        if os.path.exists(OUTPUT_FILE):
            os.remove(OUTPUT_FILE)
        return "Output file removed. <a href='/'>Go back</a>"
    except Exception as e:
        return f"Error removing file: {e}"

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.environ.get("PORT", 5000)), debug=True)
